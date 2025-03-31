"""
SAMIR IA.py: Script completo para el chatbot de Samir's Burgers
Incluye las funcionalidades de WhatsApp (Selenium y enlace directo), generación de facturas Excel
Todo integrado en un solo archivo - Optimizado para WhatsApp Web 2025
"""

import os
import sys
import json
import time
import datetime
import re
import logging
import openai
import urllib.parse
import webbrowser
from dotenv import load_dotenv
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Para WhatsApp
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys  # Añadido para usar teclas en send_message
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

# Para Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Para QR Code
try:
    import qrcode
    QR_DISPONIBLE = True
except ImportError:
    QR_DISPONIBLE = False
    print("AVISO: Para generar códigos QR, instala: pip install qrcode[pil]")

try:
    from openpyxl.drawing.image import Image  # Para el logo en Excel
except ImportError:
    print("Advertencia: No se pudo importar Image de openpyxl. El logo no se mostrará en las facturas.")
    Image = None

# Configuración de logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Cargar variables de entorno (incluye API Key de OpenAI)
load_dotenv()
openai.api_key = os.getenv('OPENAI_API_KEY')

# Constantes para la factura
EMPRESA = "Samir's Burgers"
NIT = "901.234.567-8"
DIRECCION_EMPRESA = "Calle 123 #45-67, Medellín"
TELEFONO_EMPRESA = "+57 300 123 4567"
CORREO_EMPRESA = "samirs.burgers@gmail.com"
LOGO_PATH = "logo.png"  # Ruta al logo de la empresa (opcional)

###############################
# PARTE 1: INTEGRACIÓN CON WHATSAPP
###############################

class WhatsAppBot:
    def __init__(self, data_dir=None):
        """
        Inicializa el bot de WhatsApp Web con Selenium
        
        Args:
            data_dir: Directorio para guardar la sesión de Chrome (para no escanear QR cada vez)
        """
        self.driver = None
        self.data_dir = data_dir or os.path.join(os.getcwd(), "whatsapp_session")
        
        # Asegurar que existe el directorio para la sesión
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
            
        self._setup_driver()
        
    def _setup_driver(self):
        """Configura el driver de Chrome con opciones para WhatsApp Web"""
        chrome_options = Options()
        chrome_options.add_argument(f"user-data-dir={self.data_dir}")
        
        # Opciones adicionales para mejorar la estabilidad
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-notifications")
        
        try:
            # Sintaxis actualizada para Selenium moderno
            self.driver = webdriver.Chrome(
                options=chrome_options
            )
            self.driver.maximize_window()
        except Exception as e:
            logger.error(f"Error al inicializar Chrome: {e}")
            raise
    
    def _take_screenshot(self, filename):
        """
        Toma una captura de pantalla para debug
        
        Args:
            filename: Nombre del archivo para guardar la captura
        """
        try:
            screenshot_dir = "debug_screenshots"
            if not os.path.exists(screenshot_dir):
                os.makedirs(screenshot_dir)
                
            filepath = os.path.join(screenshot_dir, filename)
            self.driver.save_screenshot(filepath)
            logger.info(f"Captura de pantalla guardada: {filepath}")
        except Exception as e:
            logger.error(f"Error al tomar captura de pantalla: {e}")
    
    def start(self):
        """Inicia WhatsApp Web y verifica la autenticación - Versión optimizada sin espera de QR"""
        logger.info("Iniciando WhatsApp Web (modo rápido)...")
        
        try:
            self.driver.get("https://web.whatsapp.com/")
            
            # Lista de posibles selectores que indican que WhatsApp Web está cargado
            possible_selectors = [
                "[data-testid='chat-list']",
                "[data-testid='default-user']",
                "[data-testid='search-tab']",
                "[data-testid='menu-bar-menu']",
                ".two",
                "._3sh5K",
                "#side",
                "#pane-side"
            ]
            
            # Tiempo máximo de espera: 15 segundos (asumiendo que ya está conectado)
            wait_time = 15
            print(f"Esperando a que cargue WhatsApp Web (máximo {wait_time} segundos)...")
            
            # Intentar cada selector
            loaded = False
            for selector in possible_selectors:
                try:
                    WebDriverWait(self.driver, wait_time).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    loaded = True
                    logger.info(f"WhatsApp Web cargado correctamente (selector: {selector})")
                    break
                except:
                    continue
            
            # Si ningún selector funcionó pero la URL cambió, considera que está cargado
            if not loaded:
                current_url = self.driver.current_url
                if "web.whatsapp.com" in current_url and not "/welcome" in current_url:
                    loaded = True
                    logger.info("WhatsApp Web parece estar cargado basado en la URL")
            
            # Si no se pudo cargar, intentar una vez más con un tiempo de espera extendido
            if not loaded:
                logger.warning("No se detectó carga de WhatsApp Web. Verificando si necesita escanear QR...")
                
                # Tomar captura para ver si muestra código QR
                self._take_screenshot("whatsapp_loading_screen.png")
                
                # Extender tiempo de espera si parece que está mostrando QR
                extended_wait = 30
                print(f"Es posible que necesites escanear el código QR. Espera {extended_wait} segundos adicionales...")
                
                time.sleep(extended_wait)
                
                # Verificar una vez más si ya está cargado
                for selector in possible_selectors:
                    try:
                        if self.driver.find_element(By.CSS_SELECTOR, selector):
                            loaded = True
                            logger.info(f"WhatsApp Web cargado después de espera extendida")
                            break
                    except:
                        continue
            
            if loaded:
                return True
            else:
                logger.error("No se pudo detectar que WhatsApp Web haya cargado correctamente")
                return False
                
        except Exception as e:
            logger.error(f"Error al iniciar WhatsApp Web: {e}")
            return False
    
    def find_chat(self, phone_number):
        """
        Busca o inicia un chat con un número específico en WhatsApp Web 2025
        
        Args:
            phone_number: Número de teléfono con código de país (ej: "573042535003")
        
        Returns:
            bool: True si se encontró/inició el chat, False en caso contrario
        """
        try:
            # Limpiar el número y asegurar que tiene el formato correcto
            clean_number = ''.join(filter(str.isdigit, phone_number))
            
            # Si no comienza con el código de país, añadirlo (Colombia = 57)
            if not clean_number.startswith('57'):
                clean_number = '57' + clean_number
                
            logger.info(f"Buscando chat para el número: {clean_number}")
            
            # Método 1: URL directa (método más confiable)
            self.driver.get(f"https://web.whatsapp.com/send?phone={clean_number}")
            logger.info("Navegando a URL directa de WhatsApp")
            
            # Esperar más tiempo (45 segundos) para cargar completamente
            try:
                # Intentar detectar panel de conversación con múltiples selectores posibles
                selectors = [
                    "[data-testid='conversation-panel-wrapper']",
                    "[data-testid='msg-container']", 
                    ".copyable-area",
                    "#main",
                    "[role='application']"
                ]
                
                for selector in selectors:
                    try:
                        # Incrementar el tiempo de espera para dar más tiempo a la carga
                        WebDriverWait(self.driver, 45).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        logger.info(f"Chat abierto correctamente usando selector: {selector}")
                        
                        # Esperar un poco más para asegurar que se cargue completamente
                        time.sleep(3)
                        return True
                    except:
                        continue
                        
                # Si llegamos aquí, ninguno de los selectores funcionó
                logger.warning("No se pudo detectar el panel de conversación con ningún selector conocido")
            except Exception as e:
                logger.warning(f"Método 1 falló: {e}")
                
            # Verificar si hay mensaje de error antes de continuar
            try:
                # Buscar textos de error comunes en la página
                error_texts = [
                    "El número de teléfono compartido a través del enlace",
                    "El número de teléfono no existe",
                    "número no está disponible",
                    "invalid",
                    "no se encuentra"
                ]
                
                for error in error_texts:
                    try:
                        # Buscar por texto parcial
                        error_element = self.driver.find_element(By.XPATH, f"//*[contains(text(), '{error}')]")
                        if error_element:
                            logger.error(f"WhatsApp indica que hay un problema con el número: {error}")
                            return False
                    except:
                        pass
            except:
                pass
            
            # Método 2: Usar el campo de búsqueda
            try:
                # Volver a la página principal
                self.driver.get("https://web.whatsapp.com/")
                logger.info("Volviendo a la página principal para intentar búsqueda manual")
                
                # Esperar a que cargue la página principal
                time.sleep(10)
                
                # Buscar el campo de búsqueda
                search_selectors = [
                    "[data-testid='chat-list-search']",
                    "[data-testid='search-bar']", 
                    "[title='Cuadro de texto de búsqueda']",
                    "._3SZ1t"
                ]
                
                search_box = None
                for selector in search_selectors:
                    try:
                        search_box = WebDriverWait(self.driver, 15).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        if search_box:
                            logger.info(f"Campo de búsqueda encontrado con selector: {selector}")
                            break
                    except:
                        continue
                
                if search_box:
                    # Limpiar y usar el número
                    search_box.clear()
                    search_box.send_keys(clean_number)
                    logger.info(f"Buscando número: {clean_number}")
                    time.sleep(5)
                    
                    # Intentar hacer clic en el resultado (si existe)
                    try:
                        chat_results = self.driver.find_elements(By.CSS_SELECTOR, "div[role='row']")
                        if chat_results and len(chat_results) > 0:
                            chat_results[0].click()
                            logger.info("Se hizo clic en el primer resultado de la búsqueda")
                            time.sleep(5)
                            return True
                    except Exception as e:
                        logger.warning(f"No se pudo hacer clic en el resultado: {e}")
                else:
                    logger.warning("No se pudo encontrar el campo de búsqueda")
            except Exception as e:
                logger.warning(f"Método 2 falló: {e}")
             
            # Método 3: Intentar buscar por el número exacto
            try:
                # Intentar encontrar directamente el chat por el número
                chat_selector = f"//span[contains(@title, '{clean_number}')]"
                chat_element = self.driver.find_element(By.XPATH, chat_selector)
                if chat_element:
                    chat_element.click()
                    logger.info(f"Chat encontrado y seleccionado por número visible")
                    time.sleep(3)
                    return True
            except:
                logger.warning("No se pudo encontrar el chat directamente por número")
                
            # Si llegamos aquí, todos los métodos fallaron
            logger.error("Todos los métodos para encontrar el chat fallaron")
            return False
        except Exception as e:
            logger.error(f"Error general al buscar chat: {e}")
            return False

    def send_message(self, phone_number, message):
        """
        Envía un mensaje a un número específico
        
        Args:
            phone_number: Número de teléfono con código de país
            message: Mensaje a enviar
            
        Returns:
            bool: True si se envió correctamente, False en caso contrario
        """
        try:
            # Primero encontrar/abrir el chat
            if not self.find_chat(phone_number):
                logger.error("No se pudo abrir el chat para enviar mensaje")
                return False
                
            # Esperar a que cargue la página completamente
            time.sleep(5)
            logger.info("Chat abierto, preparando para enviar mensaje...")
            
            # Intentar diferentes selectores para el campo de texto
            input_box = None
            input_selectors = [
                "[data-testid='conversation-compose-box-input']",
                "div[contenteditable='true'][data-tab='10']",
                "div[contenteditable='true'][role='textbox']",
                "div[role='textbox']",
                "div.selectable-text[contenteditable='true']",
                "#main div[contenteditable='true']"
            ]
            
            for selector in input_selectors:
                try:
                    input_box = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                    )
                    if input_box:
                        logger.info(f"Campo de texto encontrado con selector: {selector}")
                        break
                except:
                    continue
            
            if not input_box:
                logger.error("No se pudo encontrar el campo de texto para enviar mensaje")
                self._take_screenshot("error_send_message.png")
                return False
            
            # Enviar el mensaje de una forma más robusta
            try:
                # Método 1: Enviar directamente
                input_box.clear()
                
                # Enviar por partes para evitar errores
                chunk_size = 50  # Enviar en trozos para evitar problemas
                for i in range(0, len(message), chunk_size):
                    chunk = message[i:i+chunk_size]
                    input_box.send_keys(chunk)
                    time.sleep(0.5)
                    
                # Tomar captura del momento antes de enviar
                self._take_screenshot("before_send.png")
                logger.info("Mensaje escrito correctamente, buscando botón de enviar...")
                
                # Buscar botón de enviar con múltiples selectores
                send_button = None
                send_selectors = [
                    "[data-testid='compose-btn-send']",
                    "[data-icon='send']",
                    "[data-testid='send']",
                    "[aria-label='Enviar']",
                    "button[aria-label='Enviar']",
                    "span[data-icon='send']"
                ]
                
                for selector in send_selectors:
                    try:
                        send_button = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                        )
                        if send_button:
                            logger.info(f"Botón de enviar encontrado con selector: {selector}")
                            break
                    except:
                        continue
                
                if send_button:
                    send_button.click()
                    logger.info("Botón de enviar presionado")
                else:
                    # Alternativa: Usar tecla ENTER
                    logger.info("Usando ENTER como alternativa para enviar")
                    input_box.send_keys(Keys.ENTER)
                
                # Esperar a que se envíe el mensaje
                time.sleep(3)
                logger.info(f"Mensaje enviado correctamente a {phone_number}")
                return True
                
            except Exception as e:
                logger.error(f"Error al enviar el mensaje: {e}")
                return False
                
        except Exception as e:
            logger.error(f"Error general en send_message: {e}")
            return False
    
    def send_document(self, phone_number, file_path, caption=None):
        """
        Envía un documento (PDF, Excel, etc.) a un número específico por WhatsApp
        
        Args:
            phone_number: Número de teléfono con código de país
            file_path: Ruta absoluta al archivo para enviar
            caption: Texto opcional para el documento
            
        Returns:
            bool: True si se envió correctamente, False en caso contrario
        """
        try:
            # Asegurar que el archivo existe
            if not os.path.exists(file_path):
                logger.error(f"Archivo no encontrado: {file_path}")
                return False
                
            # Abrir chat
            if not self.find_chat(phone_number):
                logger.error("No se pudo abrir el chat para enviar documento")
                return False
            
            # Tomar captura del chat abierto
            self._take_screenshot("before_attach_document.png")
            
            # Esperar a que cargue completamente
            time.sleep(5)
                
            # PASO 1: Encontrar y hacer clic en el botón de adjuntar (clip o +)
            attach_button = None
            attach_selectors = [
                "[data-testid='attach-clip']",
                "[data-testid='compose-btn-attach']",
                "[data-icon='attach-menu-plus']",
                "[data-icon='attach']",
                "[data-icon='clip']",
                "[aria-label='Adjuntar']",
                "[title='Adjuntar']"
            ]
            
            for selector in attach_selectors:
                try:
                    attach_button = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    if attach_button:
                        logger.info(f"Botón de adjuntar encontrado con selector: {selector}")
                        break
                except:
                    continue
                    
            if not attach_button:
                logger.error("No se pudo encontrar el botón para adjuntar archivos")
                self._take_screenshot("error_no_attach_button.png")
                return False
                
            # Hacer clic en el botón de adjuntar
            attach_button.click()
            logger.info("Clic en botón de adjuntar realizado")
            time.sleep(2)
            
            # PASO 2: Buscar la opción de documento o el input de archivo directamente
            # Primero intentar encontrar y hacer clic en la opción de documento si es necesario
            document_option_selectors = [
                "[data-testid='mi-attach-document']",
                "[data-testid='attach-document']",
                "[data-icon='document']", 
                "[aria-label='Documento']"
            ]
            
            document_option_found = False
            for selector in document_option_selectors:
                try:
                    option = self.driver.find_element(By.CSS_SELECTOR, selector)
                    if option:
                        option.click()
                        logger.info(f"Opción de documento encontrada y seleccionada: {selector}")
                        document_option_found = True
                        time.sleep(2)
                        break
                except:
                    continue
            
            # PASO 3: Buscar el input de tipo file para subir el archivo
            try:
                # Tomar captura antes de buscar el input
                self._take_screenshot("before_file_input.png")
                
                file_input = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
                )
                
                # Asegurar ruta absoluta y enviar al input
                abs_file_path = os.path.abspath(file_path)
                file_input.send_keys(abs_file_path)
                logger.info(f"Archivo seleccionado: {abs_file_path}")
                
                # Esperar a que se cargue el archivo
                time.sleep(5)
                logger.info("Archivo cargado, buscando campo de comentario...")
                
                # PASO 4: Si hay caption, intentar escribirlo
                if caption:
                    try:
                        caption_selectors = [
                            "[data-testid='media-caption-input']",
                            "[data-testid='media-caption-input-container']",
                            "[data-testid='caption-input']",
                            "div[role='textbox'][data-tab='9']",
                            "[placeholder='Añade un comentario']"
                        ]
                        
                        caption_field = None
                        for selector in caption_selectors:
                            try:
                                caption_field = WebDriverWait(self.driver, 8).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                                )
                                if caption_field:
                                    logger.info(f"Campo de comentario encontrado: {selector}")
                                    break
                            except:
                                continue
                                
                        if caption_field:
                            caption_field.clear()
                            caption_field.send_keys(caption)
                            logger.info("Comentario agregado al archivo")
                            
                            # IMPORTANTE: Presionar ENTER después de escribir el comentario
                            # Esta es la nueva línea que hemos agregado para usar Enter como alternativa
                            caption_field.send_keys(Keys.ENTER)
                            logger.info("Tecla ENTER presionada después del comentario")
                            time.sleep(3)  # Esperar después de presionar Enter
                            
                            # Si el mensaje se envió con Enter, retornamos éxito
                            try:
                                # Verificar si seguimos en la pantalla de adjuntar o si ya se envió
                                if not self.driver.find_elements(By.CSS_SELECTOR, "input[type='file']"):
                                    logger.info("Documento enviado con Enter después del comentario")
                                    time.sleep(2)
                                    return True
                            except:
                                pass
                        else:
                            logger.warning("No se pudo encontrar el campo para agregar comentario")
                    except Exception as e:
                        logger.warning(f"Error al agregar comentario: {e}")
                
                # PASO 5: Buscar y hacer clic en el botón de enviar
                send_selectors = [
                    "[data-testid='send']",
                    "[data-testid='btn-send']",
                    "[aria-label='Enviar']",
                    "[data-icon='send']",
                    "span[data-icon='send']"
                ]
                
                # Tomar captura antes de buscar el botón de enviar
                self._take_screenshot("before_send_file.png")
                
                send_button = None
                for selector in send_selectors:
                    try:
                        send_button = WebDriverWait(self.driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                        )
                        if send_button:
                            logger.info(f"Botón de enviar archivo encontrado: {selector}")
                            break
                    except:
                        continue
                        
                if send_button:
                    send_button.click()
                    logger.info("Botón de enviar archivo presionado")
                else:
                    # ALTERNATIVA: Usar la tecla ENTER si no se encuentra el botón
                    logger.info("No se encontró botón de enviar, intentando con ENTER...")
                    
                    # Intentar encontrar cualquier elemento interactivo y enviar Enter
                    active_elements = [
                        "div[role='textbox']",
                        "[contenteditable='true']",
                        ".copyable-text",
                        "[data-tab='9']"
                    ]
                    
                    element_for_enter = None
                    for selector in active_elements:
                        try:
                            elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                            if elements and len(elements) > 0:
                                element_for_enter = elements[0]
                                break
                        except:
                            pass
                    
                    if element_for_enter:
                        element_for_enter.send_keys(Keys.ENTER)
                        logger.info("ENTER enviado a elemento activo")
                    else:
                        # Si no encontramos un elemento, intentar enviar Enter a la página
                        actions = webdriver.ActionChains(self.driver)
                        actions.send_keys(Keys.ENTER)
                        actions.perform()
                        logger.info("ENTER enviado a la página")
                
                # Esperar a que se envíe (archivos grandes pueden tardar más)
                time.sleep(7)
                logger.info(f"Documento enviado correctamente a {phone_number}")
                return True
                
            except Exception as e:
                logger.error(f"Error al seleccionar o enviar archivo: {e}")
                self._take_screenshot("error_file_upload.png")
                return False
                
        except Exception as e:
            logger.error(f"Error general al enviar documento: {e}")
            return False
    
    def send_image(self, phone_number, image_path, caption=None):
        """
        Envía una imagen a un número específico
        
        Args:
            phone_number: Número de teléfono con código de país
            image_path: Ruta absoluta a la imagen
            caption: Texto opcional para la imagen
            
        Returns:
            bool: True si se envió correctamente, False en caso contrario
        """
        try:
            # Abrir chat
            if not self.find_chat(phone_number):
                return False
            
            # Hacer clic en el botón de adjuntar
            attach_selectors = [
                "[data-testid='attach-clip']",
                "[data-testid='compose-btn-attach']",
                "[data-icon='attach-menu-plus']",
                "[data-icon='attach']",
                "[data-icon='clip']",
                "[aria-label='Adjuntar']",
                "[title='Adjuntar']"
            ]
            
            attach_button = None
            for selector in attach_selectors:
                try:
                    attach_button = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    if attach_button:
                        break
                except:
                    continue
                    
            if not attach_button:
                logger.error("No se pudo encontrar el botón para adjuntar")
                return False
                
            attach_button.click()
            
            # Puede ser necesario seleccionar "Imagen" si hay un menú
            image_option_selectors = [
                "[data-testid='mi-attach-gallery']",
                "[data-testid='attach-image']",
                "[data-icon='image']",
                "[aria-label='Foto o video']"
            ]
            
            # Intentar hacer clic en la opción de imagen si existe
            option_found = False
            for selector in image_option_selectors:
                try:
                    option = WebDriverWait(self.driver, 3).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    option.click()
                    option_found = True
                    time.sleep(1)
                    break
                except:
                    continue
            
            # Seleccionar opción de imagen
            image_input = WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']"))
            )
            
            # Enviar ruta de imagen al input
            image_path = os.path.abspath(image_path)  # Asegurar ruta absoluta
            image_input.send_keys(image_path)
            
            # Si hay caption, escribirlo
            if caption:
                caption_selectors = [
                    "[data-testid='media-caption-input']",
                    "[data-testid='media-caption-input-container']",
                    "[data-testid='caption-input']",
                    "div[role='textbox'][data-tab='9']",
                    "[placeholder='Añade un comentario']"
                ]
                
                caption_field = None
                for selector in caption_selectors:
                    try:
                        caption_field = WebDriverWait(self.driver, 8).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        if caption_field:
                            break
                    except:
                        continue
                        
                if caption_field:
                    caption_field.clear()
                    caption_field.send_keys(caption)
                    # Intentar con Enter
                    caption_field.send_keys(Keys.ENTER)
                    time.sleep(3)
                    
                    # Verificar si ya se envió
                    try:
                        # Si la pantalla de previsualización de imagen ya no está visible
                        if not self.driver.find_elements(By.CSS_SELECTOR, "input[type='file']"):
                            logger.info("Imagen enviada con Enter después del comentario")
                            return True
                    except:
                        pass
            
            # Hacer clic en enviar
            send_selectors = [
                "[data-testid='send']",
                "[data-testid='btn-send']",
                "[aria-label='Enviar']",
                "[data-icon='send']",
                "span[data-icon='send']"
            ]
            
            send_button = None
            for selector in send_selectors:
                try:
                    send_button = WebDriverWait(self.driver, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, selector))
                    )
                    if send_button:
                        break
                except:
                    continue
                    
            if send_button:
                send_button.click()
            else:
                # Intentar con Enter directamente en la página
                actions = webdriver.ActionChains(self.driver)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                logger.info("ENTER enviado a la página para enviar imagen")
            
            # Esperar a que se envíe
            time.sleep(3)
            logger.info(f"Imagen enviada correctamente a {phone_number}")
            return True
        except Exception as e:
            logger.error(f"Error al enviar imagen a {phone_number}: {e}")
            return False
    
    def close(self):
        """Cierra el navegador y finaliza la sesión"""
        if self.driver:
            self.driver.quit()
            logger.info("Sesión de WhatsApp cerrada")

def send_whatsapp_message(phone_number, message):
    """
    Función de utilidad para enviar un mensaje de WhatsApp sin manejar el objeto bot
    
    Args:
        phone_number: Número con código de país (ej: "573001234567")
        message: Texto del mensaje a enviar
        
    Returns:
        bool: True si se envió correctamente
    """
    bot = WhatsAppBot()
    success = False
    
    try:
        if bot.start():
            success = bot.send_message(phone_number, message)
    finally:
        bot.close()
        
    return success

def mensaje_whatsapp(datos_cliente, resumen_pedido):
    """Genera el mensaje estándar para WhatsApp"""
    return (
        f"🍔 *SAMIR'S BURGERS - CONFIRMACIÓN DE PEDIDO* 🍔\n\n"
        f"Hola *{datos_cliente['nombre']}*, ¡gracias por tu pedido!\n\n"
        f"*📝 RESUMEN DE TU PEDIDO:*\n{resumen_pedido}\n\n"
        f"*📍 DIRECCIÓN DE ENTREGA:*\n{datos_cliente['direccion']}\n\n"
        f"*💰 MÉTODO DE PAGO:*\n{datos_cliente['metodo_pago']}\n\n"
        f"Tu pedido está siendo preparado con todo el amor de Samir's Burgers. "
        f"Si tienes alguna duda, responde a este mensaje.\n\n"
        f"¡Buen provecho! 🍔✨"
    )

def enviar_pedido_por_whatsapp(telefono, resumen_pedido, datos_cliente):
    """
    Envía el resumen del pedido a un cliente por WhatsApp usando Selenium
    
    Args:
        telefono: Número de teléfono del cliente (se le añadirá el código de país)
        resumen_pedido: Texto con el resumen del pedido
        datos_cliente: Diccionario con los datos del cliente
        
    Returns:
        bool: True si se envió correctamente
    """
    # Formatear el número con código de país (Colombia por defecto)
    # Eliminar cualquier formato (espacios, guiones, +)
    clean_number = ''.join(filter(str.isdigit, telefono))
    
    # Si no comienza con el código de país, añadirlo
    if not clean_number.startswith('57'):
        clean_number = '57' + clean_number
    
    # Crear mensaje con formato
    mensaje = mensaje_whatsapp(datos_cliente, resumen_pedido)
    
    # Enviar mensaje
    return send_whatsapp_message(clean_number, mensaje)

def enviar_factura_por_whatsapp(telefono, ruta_factura, datos_cliente, resumen_pedido):
    """
    Envía la factura Excel por WhatsApp
    
    Args:
        telefono: Número de teléfono del cliente
        ruta_factura: Ruta al archivo Excel de factura
        datos_cliente: Datos del cliente
        resumen_pedido: Texto del resumen del pedido
        
    Returns:
        bool: True si se envió correctamente
    """
    # Formatear el número
    clean_number = ''.join(filter(str.isdigit, telefono))
    if not clean_number.startswith('57'):
        clean_number = '57' + clean_number
    
    # Mensaje de factura
    mensaje = (
        f"🧾 *FACTURA SAMIR'S BURGERS* 🧾\n\n"
        f"Hola *{datos_cliente['nombre']}*, a continuación te enviamos la factura de tu pedido.\n\n"
        f"Si tienes alguna duda sobre tu factura, contáctanos al {TELEFONO_EMPRESA}.\n\n"
        f"¡Gracias por tu compra!"
    )
    
    # Crear instancia del bot
    bot = WhatsAppBot()
    success = False
    
    try:
        if bot.start():
            # Primero enviar un mensaje
            bot.send_message(clean_number, mensaje)
            
            # Luego enviar el archivo Excel
            caption = f"Factura Samir's Burgers - {datetime.datetime.now().strftime('%d/%m/%Y')}"
            success = bot.send_document(clean_number, ruta_factura, caption)
    except Exception as e:
        logger.error(f"Error al enviar factura por WhatsApp: {e}")
        success = False
    finally:
        bot.close()
    
    return success

###############################
# PARTE 1B: WHATSAPP DIRECTO (QR y Enlace)
###############################

def crear_enlace_whatsapp(numero, mensaje):
    """
    Crea un enlace de WhatsApp que abre directamente el chat con un mensaje predefinido
    
    Args:
        numero: Número de teléfono (con código de país)
        mensaje: Mensaje a enviar (será codificado para URL)
        
    Returns:
        str: URL de WhatsApp
    """
    # Formatear el número
    numero_limpio = ''.join(filter(str.isdigit, numero))
    
    # Si no comienza con el código de país, añadirlo (Colombia = 57)
    if not numero_limpio.startswith('57'):
        numero_limpio = '57' + numero_limpio
    
    # Codificar el mensaje para URL
    mensaje_codificado = urllib.parse.quote(mensaje)
    
    # Crear el enlace (funciona tanto en móvil como en escritorio)
    return f"https://wa.me/{numero_limpio}?text={mensaje_codificado}"

def generar_qr_whatsapp(numero, mensaje, guardar=True):
    """
    Genera un código QR que al escanearse abre WhatsApp con un mensaje predefinido
    
    Args:
        numero: Número de teléfono (con código de país)
        mensaje: Mensaje a enviar 
        guardar: Si es True, guarda el QR como imagen
        
    Returns:
        str: Ruta a la imagen QR generada o None si no se guardó
    """
    if not QR_DISPONIBLE:
        print("Error: Para generar códigos QR, instala: pip install qrcode[pil]")
        return None
        
    try:
        # Crear el enlace
        enlace = crear_enlace_whatsapp(numero, mensaje)
        
        # Generar QR
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(enlace)
        qr.make(fit=True)
        
        # Crear imagen
        img = qr.make_image(fill_color="black", back_color="white")
        
        if guardar:
            # Crear nombre de archivo con timestamp
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"whatsapp_qr_{timestamp}.png"
            img.save(filename)
            print(f"Código QR guardado como: {filename}")
            return filename
        else:
            return None
    except Exception as e:
        print(f"Error al generar código QR: {e}")
        return None

def abrir_chat_directo(numero, mensaje):
    """
    Abre directamente el navegador predeterminado con el enlace de WhatsApp
    
    Args:
        numero: Número de teléfono (con código de país)
        mensaje: Mensaje a enviar
        
    Returns:
        bool: True si se pudo abrir el navegador
    """
    try:
        enlace = crear_enlace_whatsapp(numero, mensaje)
        webbrowser.open(enlace)
        print(f"Enlace de WhatsApp abierto: {enlace}")
        return True
    except Exception as e:
        print(f"Error al abrir enlace de WhatsApp: {e}")
        return False

def enviar_pedido_whatsapp_directo(telefono, resumen_pedido, datos_cliente):
    """
    Crea un código QR y un enlace directo para enviar el pedido por WhatsApp
    
    Args:
        telefono: Número de teléfono del cliente
        resumen_pedido: Texto con el resumen del pedido
        datos_cliente: Diccionario con los datos del cliente
        
    Returns:
        dict: Diccionario con enlace, ruta del QR y éxito de apertura
    """
    # Crear mensaje formateado
    mensaje = mensaje_whatsapp(datos_cliente, resumen_pedido)
    
    # Crear el enlace
    enlace = crear_enlace_whatsapp(telefono, mensaje)
    
    # Generar código QR
    qr_path = generar_qr_whatsapp(telefono, mensaje) if QR_DISPONIBLE else None
    
    # Preguntar si desea abrir el navegador
    print("\n¿Deseas abrir el enlace de WhatsApp en tu navegador? (s/n): ", end="")
    respuesta = input().strip().lower()
    
    browser_opened = False
    if respuesta == 's' or respuesta == 'si' or respuesta == 'y' or respuesta == 'yes':
        browser_opened = abrir_chat_directo(telefono, mensaje)
    
    # Mostrar instrucciones
    print("\n📱 INSTRUCCIONES PARA ENVIAR EL PEDIDO POR WHATSAPP:")
    if qr_path:
        print(f"1. OPCIÓN 1: Escanea el código QR generado ({qr_path})")
    print(f"2. OPCIÓN 2: Abre este enlace manualmente:")
    print(f"   {enlace}")
    print("3. El chat de WhatsApp se abrirá con el mensaje listo para enviar")
    print("4. Solo haz clic en el botón de ENVIAR en WhatsApp\n")
    
    return {
        'enlace': enlace,
        'qr_path': qr_path,
        'browser_opened': browser_opened
    }

###############################
# PARTE 2: FACTURAS EXCEL
###############################

class FacturaExcel:
    def __init__(self):
        """Inicializa el generador de facturas en Excel"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Factura"
        
        # Cargar contador de facturas
        self.contador_file = "invoice_count.json"
        self.num_factura = self._obtener_numero_factura()
        
        # Estilos predefinidos
        self.titulo_font = Font(name='Arial', size=16, bold=True)
        self.subtitulo_font = Font(name='Arial', size=12, bold=True)
        self.header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        self.normal_font = Font(name='Arial', size=11)
        
        # Colores
        self.color_principal = "FF8C00"  # Naranja para mantener la identidad de hamburguesas
        self.color_secundario = "FFC04D"  # Naranja claro
        
        # Bordes
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def _obtener_numero_factura(self):
        """
        Obtiene y actualiza el número de factura secuencial
        """
        today = datetime.date.today().isoformat()
        try:
            with open(self.contador_file, "r") as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            data = {"last_number": 0, "date": today}
        
        # Si es un nuevo día, reiniciar contador
        if data.get("date") != today:
            count = 1
        else:
            count = data.get("last_number", 0) + 1
            
        # Guardar nuevo contador
        data["last_number"] = count
        data["date"] = today
        
        with open(self.contador_file, "w") as f:
            json.dump(data, f)
            
        # Formato: YYYYMMDD-NNN
        fecha_formato = today.replace("-", "")
        return f"{fecha_formato}-{count:03d}"
        
    def _configurar_pagina(self):
        """Configura las dimensiones y márgenes de la página"""
        # Ajustar anchos de columna
        self.ws.column_dimensions['A'].width = 5   # Núm ítem
        self.ws.column_dimensions['B'].width = 40  # Descripción
        self.ws.column_dimensions['C'].width = 12  # Cantidad
        self.ws.column_dimensions['D'].width = 15  # Precio unitario
        self.ws.column_dimensions['E'].width = 15  # Total
    
    def _crear_encabezado(self):
        """Crea el encabezado de la factura con logo y datos de la empresa"""
        # Intentar insertar logo si existe
        try:
            if Image is not None and os.path.exists(LOGO_PATH):
                logo = Image(LOGO_PATH)
                logo.width = 150
                logo.height = 70
                self.ws.add_image(logo, "A1")
        except Exception as e:
            logger.warning(f"No se pudo cargar el logo: {e}")
            self.ws['A1'] = EMPRESA
            self.ws['A1'].font = Font(name='Arial', size=20, bold=True)
        
        # Información de la empresa - Lado izquierdo
        self.ws['B1'] = EMPRESA
        self.ws['B1'].font = self.titulo_font
        
        self.ws['B2'] = f"NIT: {NIT}"
        self.ws['B2'].font = self.normal_font
        
        self.ws['B3'] = DIRECCION_EMPRESA
        self.ws['B3'].font = self.normal_font
        
        self.ws['B4'] = f"Tel: {TELEFONO_EMPRESA}"
        self.ws['B4'].font = self.normal_font
        
        self.ws['B5'] = f"Email: {CORREO_EMPRESA}"
        self.ws['B5'].font = self.normal_font
        
        # Información de factura - Lado derecho
        self.ws['D1'] = "FACTURA DE VENTA"
        self.ws['D1'].font = self.titulo_font
        self.ws['D1'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('D1:E1')
        
        self.ws['D2'] = f"No. {self.num_factura}"
        self.ws['D2'].font = self.subtitulo_font
        self.ws['D2'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells('D2:E2')
        
        # Fecha de emisión
        fecha_actual = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
        self.ws['D3'] = "Fecha de Emisión:"
        self.ws['D3'].font = self.normal_font
        self.ws['D3'].alignment = Alignment(horizontal='right')
        
        self.ws['E3'] = fecha_actual
        self.ws['E3'].font = self.normal_font
        
        # Separador
        for col in range(1, 6):  # A-E
            self.ws.cell(row=7, column=col).fill = PatternFill(
                start_color=self.color_principal, 
                end_color=self.color_principal,
                fill_type="solid"
            )
    
    def _agregar_datos_cliente(self, datos_cliente):
        """
        Agrega la sección de datos del cliente
        
        Args:
            datos_cliente: Diccionario con nombre, direccion, telefono
        """
        row = 8
        
        self.ws[f'A{row}'] = "DATOS DEL CLIENTE"
        self.ws[f'A{row}'].font = self.subtitulo_font
        self.ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        self.ws[f'A{row}'] = "Nombre:"
        self.ws[f'A{row}'].font = self.header_font
        self.ws[f'A{row}'].fill = PatternFill(start_color=self.color_principal, 
                                              end_color=self.color_principal,
                                              fill_type="solid")
        
        self.ws[f'B{row}'] = datos_cliente['nombre']
        self.ws[f'B{row}'].font = self.normal_font
        self.ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        self.ws[f'A{row}'] = "Teléfono:"
        self.ws[f'A{row}'].font = self.header_font
        self.ws[f'A{row}'].fill = PatternFill(start_color=self.color_principal, 
                                              end_color=self.color_principal,
                                              fill_type="solid")
        
        self.ws[f'B{row}'] = datos_cliente['telefono']
        self.ws[f'B{row}'].font = self.normal_font
        self.ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        self.ws[f'A{row}'] = "Dirección:"
        self.ws[f'A{row}'].font = self.header_font
        self.ws[f'A{row}'].fill = PatternFill(start_color=self.color_principal, 
                                              end_color=self.color_principal,
                                              fill_type="solid")
        
        self.ws[f'B{row}'] = datos_cliente['direccion']
        self.ws[f'B{row}'].font = self.normal_font
        self.ws.merge_cells(f'B{row}:E{row}')
        
        row += 1
        self.ws[f'A{row}'] = "Pago:"
        self.ws[f'A{row}'].font = self.header_font
        self.ws[f'A{row}'].fill = PatternFill(start_color=self.color_principal, 
                                              end_color=self.color_principal,
                                              fill_type="solid")
        
        self.ws[f'B{row}'] = datos_cliente['metodo_pago']
        self.ws[f'B{row}'].font = self.normal_font
        self.ws.merge_cells(f'B{row}:E{row}')
        
        # Espacio
        row += 2
        return row
    
    def _crear_tabla_productos(self, row_start, items):
        """
        Crea la tabla de productos del pedido
        
        Args:
            row_start: Fila donde inicia la tabla
            items: Lista de diccionarios con los items del pedido
                   [{'descripcion': '', 'cantidad': n, 'precio': n, 'total': n}, ...]
                   
        Returns:
            int: Fila siguiente después de la tabla
        """
        # Encabezados de tabla
        headers = ["#", "Descripción", "Cantidad", "Precio Unit.", "Total"]
        
        row = row_start
        self.ws[f'A{row}'] = "DETALLE DEL PEDIDO"
        self.ws[f'A{row}'].font = self.subtitulo_font
        self.ws.merge_cells(f'A{row}:E{row}')
        
        # Cabecera de tabla
        row += 1
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            self.ws[f'{col_letter}{row}'] = header
            self.ws[f'{col_letter}{row}'].font = self.header_font
            self.ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')
            self.ws[f'{col_letter}{row}'].fill = PatternFill(
                start_color=self.color_principal, 
                end_color=self.color_principal,
                fill_type="solid"
            )
            # Borde
            self.ws[f'{col_letter}{row}'].border = self.thin_border
        
        # Contenido de la tabla
        total_general = 0
        domicilio = 0
        
        for i, item in enumerate(items, 1):
            row += 1
            
            # Número de ítem
            self.ws[f'A{row}'] = i
            self.ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            # Descripción
            self.ws[f'B{row}'] = item['descripcion']
            
            # Cantidad
            self.ws[f'C{row}'] = item['cantidad']
            self.ws[f'C{row}'].alignment = Alignment(horizontal='center')
            
            # Precio unitario
            self.ws[f'D{row}'] = item['precio']
            self.ws[f'D{row}'].number_format = '"$"#,##0'
            self.ws[f'D{row}'].alignment = Alignment(horizontal='right')
            
            # Total
            total_item = item['cantidad'] * item['precio']
            self.ws[f'E{row}'] = total_item
            self.ws[f'E{row}'].number_format = '"$"#,##0'
            self.ws[f'E{row}'].alignment = Alignment(horizontal='right')
            
            # Si es domicilio, guardarlo aparte
            if 'domicilio' in item['descripcion'].lower():
                domicilio = total_item
            else:
                total_general += total_item
            
            # Bordes para toda la fila
            for col_idx in range(1, 6):
                col_letter = get_column_letter(col_idx)
                self.ws[f'{col_letter}{row}'].border = self.thin_border
                
        # Subtotal y domicilio
        row += 2
        self.ws[f'D{row}'] = "Subtotal:"
        self.ws[f'D{row}'].font = self.header_font
        self.ws[f'D{row}'].alignment = Alignment(horizontal='right')
        
        self.ws[f'E{row}'] = total_general
        self.ws[f'E{row}'].number_format = '"$"#,##0'
        self.ws[f'E{row}'].alignment = Alignment(horizontal='right')
        self.ws[f'E{row}'].font = self.normal_font
        
        row += 1
        self.ws[f'D{row}'] = "Domicilio:"
        self.ws[f'D{row}'].font = self.header_font
        self.ws[f'D{row}'].alignment = Alignment(horizontal='right')
        
        self.ws[f'E{row}'] = domicilio
        self.ws[f'E{row}'].number_format = '"$"#,##0'
        self.ws[f'E{row}'].alignment = Alignment(horizontal='right')
        self.ws[f'E{row}'].font = self.normal_font
        
        row += 1
        self.ws[f'D{row}'] = "TOTAL A PAGAR:"
        self.ws[f'D{row}'].font = self.header_font
        self.ws[f'D{row}'].alignment = Alignment(horizontal='right')
        
        self.ws[f'E{row}'] = total_general + domicilio
        self.ws[f'E{row}'].number_format = '"$"#,##0'
        self.ws[f'E{row}'].alignment = Alignment(horizontal='right')
        self.ws[f'E{row}'].font = self.subtitulo_font
        
        # Agregar notas o términos
        row += 3
        self.ws[f'A{row}'] = "NOTAS:"
        self.ws[f'A{row}'].font = self.subtitulo_font
        self.ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        self.ws[f'A{row}'] = "• Esta factura es un comprobante válido para reclamaciones y garantías."
        self.ws[f'A{row}'].font = self.normal_font
        self.ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        self.ws[f'A{row}'] = "• Cualquier inconveniente con su pedido, contáctenos al " + TELEFONO_EMPRESA
        self.ws[f'A{row}'].font = self.normal_font
        self.ws.merge_cells(f'A{row}:E{row}')
        
        row += 2
        self.ws[f'A{row}'] = "¡GRACIAS POR SU COMPRA!"
        self.ws[f'A{row}'].font = self.subtitulo_font
        self.ws[f'A{row}'].alignment = Alignment(horizontal='center')
        self.ws.merge_cells(f'A{row}:E{row}')
        
        return row + 1
        
    def generar_factura(self, datos_cliente, resumen_pedido):
        """
        Genera una factura completa en Excel
        
        Args:
            datos_cliente: Diccionario con datos del cliente
            resumen_pedido: Texto con el resumen del pedido
            
        Returns:
            str: Ruta al archivo Excel generado
        """
        # Extraer items del pedido desde el resumen de texto
        items = self._extraer_items_del_resumen(resumen_pedido)
        
        # Configurar página
        self._configurar_pagina()
        
        # Crear encabezado
        self._crear_encabezado()
        
        # Agregar datos del cliente
        row = self._agregar_datos_cliente(datos_cliente)
        
        # Agregar tabla de productos
        self._crear_tabla_productos(row, items)
        
        # Guardar archivo
        fecha = datetime.date.today().strftime("%Y%m%d")
        nombre_archivo = f"Factura_{self.num_factura}_{datos_cliente['nombre'].replace(' ', '_')}.xlsx"
        
        ruta_archivo = os.path.join(os.getcwd(), nombre_archivo)
        self.wb.save(ruta_archivo)
        
        logger.info(f"Factura generada: {ruta_archivo}")
        return ruta_archivo
        
    def _extraer_items_del_resumen(self, resumen_pedido):
        """
        Extrae los items y precios del resumen textual del pedido
        
        Args:
            resumen_pedido: Texto con el resumen del pedido
            
        Returns:
            list: Lista de diccionarios con detalles de items
        """
        items = []
        
        # Dividir por líneas y buscar patrones de items y precios
        lineas = resumen_pedido.strip().split('\n')
        
        for linea in lineas:
            linea = linea.strip()
            if not linea:
                continue
            
            # Algunos patrones comunes:
            # 1. "X Hamburguesa Clásica - $12000"
            # 2. "Hamburguesa Clásica (x2) - $24000"
            # 3. "Domicilio - $2000"
            
            # Patrón 1: "X Item - $Precio"
            patron1 = r'(\d+)\s+(.+?)\s*-\s*\$?(\d+(?:,\d+)?)'
            coincidencia = re.search(patron1, linea)
            
            if coincidencia:
                cantidad = int(coincidencia.group(1))
                descripcion = coincidencia.group(2).strip()
                precio_total = float(coincidencia.group(3).replace(',', ''))
                precio_unitario = precio_total / cantidad
                
                items.append({
                    'descripcion': descripcion,
                    'cantidad': cantidad,
                    'precio': precio_unitario,
                    'total': precio_total
                })
                continue
            
            # Patrón 2: "Item (xX) - $Precio"
            patron2 = r'(.+?)\s*\(x(\d+)\)\s*-\s*\$?(\d+(?:,\d+)?)'
            coincidencia = re.search(patron2, linea)
            
            if coincidencia:
                descripcion = coincidencia.group(1).strip()
                cantidad = int(coincidencia.group(2))
                precio_total = float(coincidencia.group(3).replace(',', ''))
                precio_unitario = precio_total / cantidad
                
                items.append({
                    'descripcion': descripcion,
                    'cantidad': cantidad,
                    'precio': precio_unitario,
                    'total': precio_total
                })
                continue
            
            # Patrón 3: "Item - $Precio" (cantidad 1 implícita)
            patron3 = r'(.+?)\s*-\s*\$?(\d+(?:,\d+)?)'
            coincidencia = re.search(patron3, linea)
            
            if coincidencia:
                descripcion = coincidencia.group(1).strip()
                precio = float(coincidencia.group(2).replace(',', ''))
                
                # Si contiene "domicilio" o términos similares
                if any(term in descripcion.lower() for term in ["domicilio", "envío", "delivery"]):
                    items.append({
                        'descripcion': "Domicilio",
                        'cantidad': 1,
                        'precio': precio,
                        'total': precio
                    })
                else:
                    items.append({
                        'descripcion': descripcion,
                        'cantidad': 1,
                        'precio': precio,
                        'total': precio
                    })
                continue
        
        # Si no se identificaron items, crear uno genérico
        if not items:
            items.append({
                'descripcion': "Pedido completo (ver detalle en resumen)",
                'cantidad': 1,
                'precio': 0,  # Precio desconocido
                'total': 0
            })
        
        return items

def generar_y_enviar_factura(datos_cliente, resumen_pedido):
    """
    Genera una factura Excel y la envía por correo
    
    Args:
        datos_cliente: Diccionario con datos del cliente
        resumen_pedido: Texto del resumen del pedido
        
    Returns:
        tuple: (éxito_generación, éxito_envío, ruta_factura)
    """
    try:
        # Crear generador de facturas
        factura_gen = FacturaExcel()
        
        # Generar factura
        ruta_factura = factura_gen.generar_factura(datos_cliente, resumen_pedido)
        
        # Enviar por correo
        enviado = enviar_factura_por_correo(ruta_factura, datos_cliente, resumen_pedido)
        
        return True, enviado, ruta_factura
    
    except Exception as e:
        logger.error(f"Error al generar/enviar factura: {e}")
        return False, False, None

def enviar_factura_por_correo(ruta_factura, datos_cliente, resumen_pedido):
    """
    Envía la factura por correo electrónico
    
    Args:
        ruta_factura: Ruta al archivo Excel de la factura
        datos_cliente: Diccionario con datos del cliente
        resumen_pedido: Texto con el resumen del pedido
        
    Returns:
        bool: True si se envió correctamente
    """
    # Datos de correo electrónico
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    smtp_user = "apeironiafilo@gmail.com"  # Usar el mismo del código original
    smtp_password = "nuhk ijhw szaf rgut"  # Usar el mismo del código original
    
    # Destinatarios
    to_email = "samirosorio21@gmail.com , josedanielorregor@gmail.com , haidyflorez98@gmail.com , ab4962267@gmail.com "  # Usar los mismos del código original
    
    try:
        # Crear mensaje
        mensaje = MIMEMultipart()
        mensaje['From'] = f"Samir's Burgers <{smtp_user}>"
        mensaje['To'] = to_email
        mensaje['Subject'] = f"🧾 Factura - Samir's Burgers - Pedido {os.path.basename(ruta_factura).split('_')[1]}"
        
        # Cuerpo HTML del correo
        html = f"""
        <html>
          <body style="font-family: Arial, sans-serif;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 5px;">
              <div style="text-align: center; margin-bottom: 20px;">
                <h1 style="color: #FF8C00;">🧾 Factura - Samir's Burgers</h1>
              </div>
              
              <p>Estimado cliente,</p>
              
              <p>Adjunto encontrará la <strong>factura electrónica</strong> correspondiente a su pedido en Samir's Burgers.</p>
              
              <div style="background-color: #f9f9f9; padding: 15px; border-radius: 5px; margin: 15px 0;">
                <h3 style="margin-top: 0; color: #FF8C00;">Resumen del Pedido:</h3>
                <p style="white-space: pre-line;">{resumen_pedido}</p>
              </div>
              
              <p>La factura adjunta es un documento válido para efectos fiscales y garantías de servicio.</p>
              
              <p>Si tiene alguna pregunta o inquietud, no dude en contactarnos al {TELEFONO_EMPRESA}.</p>
              
              <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; text-align: center; color: #777;">
                <p>¡Gracias por su preferencia!</p>
                <p>Samir's Burgers - Sabor que enamora</p>
                <p>{DIRECCION_EMPRESA}</p>
              </div>
            </div>
          </body>
        </html>
        """
        
        # Adjuntar HTML
        mensaje.attach(MIMEText(html, "html"))
        
        # Adjuntar factura
        with open(ruta_factura, "rb") as archivo:
            adjunto = MIMEBase("application", "octet-stream")
            adjunto.set_payload(archivo.read())
        
        encoders.encode_base64(adjunto)
        adjunto.add_header(
            "Content-Disposition", 
            f"attachment; filename={os.path.basename(ruta_factura)}"
        )
        mensaje.attach(adjunto)
        
        # Enviar correo
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(mensaje)
            
        logger.info(f"Factura enviada exitosamente por correo a {to_email}")
        return True
    
    except Exception as e:
        logger.error(f"Error al enviar factura por correo: {e}")
        return False

###############################
# PARTE 3: FUNCIONES PRINCIPALES DEL CHATBOT
###############################

def actualizar_contador_pedidos():
    """
    Actualiza y retorna el número de pedido para el día actual.
    Se almacena en un archivo 'order_count.json' que contiene un diccionario
    con la fecha (YYYY-MM-DD) como clave y el contador como valor.
    """
    contador_file = "order_count.json"
    today = datetime.date.today().isoformat()
    try:
        with open(contador_file, "r") as f:
            data = json.load(f)
    except FileNotFoundError:
        data = {}
    count = data.get(today, 0) + 1
    data[today] = count
    with open(contador_file, "w") as f:
        json.dump(data, f)
    return count

def obtener_respuesta(conversacion):
    """Envía la conversación a OpenAI y retorna la respuesta."""
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4o",  # Usa un modelo válido
            messages=conversacion,
            max_tokens=1000
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print("Error al obtener respuesta de OpenAI:", e)
        return "😕 Lo siento, ha ocurrido un error."

def generar_resumen_pedido(conversacion, datos_cliente):
    """
    Genera un resumen completo del pedido combinando la conversación y los datos del cliente.
    """
    prompt = (
        "A partir de la siguiente conversación, genera un resumen completo del pedido, "
        "incluyendo todos los ítems, combos, promociones y detalles de domicilio. "
        "Asegúrate de incluir las cantidades exactas y los precios de cada ítem. "
        "El formato debe ser: X [nombre del producto] - $[precio]. "
        "Luego, añade la siguiente información del cliente:\n\n"
    )
    prompt += (
        f"Nombre: {datos_cliente['nombre']}\n"
        f"Teléfono: {datos_cliente['telefono']}\n"
        f"Dirección: {datos_cliente['direccion']}\n"
        f"Método de pago: {datos_cliente['metodo_pago']}\n\n"
    )
    prompt += "Conversación:\n"
    for msg in conversacion:
        prompt += f"{msg['role']}: {msg['content']}\n"
    prompt += "\nResumen del pedido:"
    
    try:
        resumen_response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "Eres un asistente que resume pedidos de un restaurante de forma clara y detallada. Asegúrate de especificar las cantidades y los precios de cada ítem."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=500
        )
        resumen = resumen_response.choices[0].message.content.strip()
        return resumen
    except Exception as e:
        print("Error al generar el resumen del pedido:", e)
        return "❌ Error al generar el resumen del pedido."

def formatear_pedido_en_bullets(pedido):
    """
    Toma la cadena 'pedido' y la convierte en una lista HTML con viñetas.
    Se asume que cada ítem está separado por saltos de línea.
    """
    items = pedido.splitlines()
    bullet_list = "<ul>"
    for item in items:
        if item.strip():
            bullet_list += f"<li>{item.strip()}</li>"
    bullet_list += "</ul>"
    return bullet_list

def enviar_pedido_por_correo(pedido, datos_cliente):
    """
    Envía el resumen del pedido por correo electrónico con formato HTML y un asunto
    que incluye el número del pedido (contador diario).
    Se organiza en secciones para el coordinador, domiciliario y contable.
    """
    # Actualizar el contador y obtener el número de pedido del día
    count = actualizar_contador_pedidos()
    
    # Formatear el contenido del pedido en bullets
    bullet_pedido = formatear_pedido_en_bullets(pedido)
    
    cuerpo = f"""
    <html>
      <body style="font-family: Arial, sans-serif;">
        <hr>
        <h2 style="text-align: center;">🍔 PEDIDO FINAL - Samir's Burgers 🍔</h2>
        <h3 style="text-align: center;">Pedido #{count}</h3>
        <hr>
        <h3>👨‍🍳 Pedido:</h3>
        {bullet_pedido}
        <hr>
        <h3>📞 Información del Cliente:</h3>
        <ul>
          <li><strong>Nombre:</strong> {datos_cliente['nombre']}</li>
          <li><strong>Teléfono:</strong> {datos_cliente['telefono']}</li>
          <li><strong>Dirección de entrega:</strong> {datos_cliente['direccion']}</li>
          <li><strong>Método de pago:</strong> {datos_cliente['metodo_pago']}</li>
        </ul>
        <hr>
        <h3>📝 Notas para el Coordinador:</h3>
        <p>Verificar la exactitud del pedido y confirmar la preparación.</p>
        <h3>🚗 Notas para el Domiciliario:</h3>
        <p>Entregar el pedido a la dirección indicada y contactar al cliente al llegar.</p>
        <h3>💰 Notas para el Contable:</h3>
        <p>Registrar el total a pagar y facturar el pedido.</p>
        <hr>
        <p style="text-align: center;">¡Gracias por tu preferencia! 🎉</p>
      </body>
    </html>
    """
    msg = MIMEText(cuerpo, "html", "utf-8")
    msg['Subject'] = f"🍔 Pedido Final - Samir's Burgers - Pedido #{count} 🍔"
    msg['From'] = "apeironiafilo@gmail.com"       # Tu correo de envío
    msg['To'] = "jmct_74@gmail.com, danielalvarezmoncada527@gmail.com"             # Correo de destino
     # Correo de destino

    smtp_server = "smtp.gmail.com"
    smtp_port = 587  # TLS
    smtp_user = "apeironiafilo@gmail.com"
    smtp_password = "nuhk ijhw szaf rgut"  # Contraseña de aplicación

    try:
        print("[DEBUG] Conectando al servidor SMTP...")
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            print("[DEBUG] Iniciando sesión en SMTP...")
            server.login(smtp_user, smtp_password)
            print("[DEBUG] Enviando correo...")
            server.send_message(msg)
        print("✅ Pedido enviado exitosamente por correo.")
        return True
    except Exception as e:
        print("❌ Error al enviar el correo:", e)
        return False

def procesar_opciones_envio(datos_cliente, resumen_pedido):
    """
    Procesa las opciones de envío del pedido: correo HTML, factura Excel y WhatsApp
    
    Args:
        datos_cliente: Diccionario con los datos del cliente
        resumen_pedido: Texto con el resumen del pedido
    """
    opciones_validas = ['1', '2', '3', '4', '5', '6', '7']
    opciones_seleccionadas = []
    
    while True:
        print("\n📱 OPCIONES DE ENVÍO DEL PEDIDO:")
        print("1. Enviar resumen por correo (formato HTML)")
        print("2. Enviar factura Excel por correo")
        print("3. Enviar confirmación por WhatsApp (mensaje)")
        print("4. Enviar confirmación por WhatsApp (enlace directo/QR)")
        print("5. Enviar todas las opciones de correo (1 y 2)")
        print("6. Enviar TODAS las opciones anteriores (completo)")
        print("7. Finalizar sin enviar")
        
        seleccion = input("\nSelecciona una opción (1-7): ").strip()
        
        if seleccion not in opciones_validas:
            print("❌ Opción no válida. Intenta nuevamente.")
            continue
        
        # Salir del bucle si la opción es válida
        opciones_seleccionadas.append(seleccion)
        break
    
    resultados = {
        'email_html': False,
        'email_excel': False,
        'whatsapp': False,
        'whatsapp_direct': False,
        'whatsapp_file': False
    }
    
    # Variables para factura
    ruta_factura = None
    exito_gen_factura = False
    
    # Procesar opciones seleccionadas
    if '6' in opciones_seleccionadas:  # TODAS las opciones
        print("\n🚀 Procesando TODOS los métodos de envío...")
        
        # 1. Enviar resumen por correo (HTML)
        print("\n📧 Enviando resumen por correo (HTML)...")
        resultados['email_html'] = enviar_pedido_por_correo(resumen_pedido, datos_cliente)
        
        # 2. Generar factura Excel (la necesitamos para WhatsApp también)
        print("\n📊 Generando factura Excel...")
        exito_gen_factura, exito_envio, ruta_factura = generar_y_enviar_factura(datos_cliente, resumen_pedido)
        resultados['email_excel'] = exito_envio
        
        # 3. Enviar mensaje por WhatsApp (Selenium)
        print("\n📱 Enviando mensaje por WhatsApp (Selenium)...")
        resultados['whatsapp'] = enviar_pedido_por_whatsapp(datos_cliente['telefono'], resumen_pedido, datos_cliente)
        
        # 3b. Enviar factura por WhatsApp si se generó correctamente
        if exito_gen_factura and ruta_factura:
            print("\n📎 Enviando factura Excel por WhatsApp...")
            resultados['whatsapp_file'] = enviar_factura_por_whatsapp(
                datos_cliente['telefono'], 
                ruta_factura, 
                datos_cliente, 
                resumen_pedido
            )
        
        # 4. Generar enlace y QR para WhatsApp directo
        print("\n🔗 Generando enlace y QR para WhatsApp directo...")
        try:
            if QR_DISPONIBLE:
                resultado_direct = enviar_pedido_whatsapp_directo(
                    datos_cliente['telefono'], 
                    resumen_pedido, 
                    datos_cliente
                )
                resultados['whatsapp_direct'] = True
            else:
                print("Para generar QR, instala: pip install qrcode[pil]")
                enlace = crear_enlace_whatsapp(datos_cliente['telefono'], mensaje_whatsapp(datos_cliente, resumen_pedido))
                print(f"Enlace directo: {enlace}")
                resultados['whatsapp_direct'] = False
        except Exception as e:
            print(f"❌ Error al generar enlace de WhatsApp: {e}")
            resultados['whatsapp_direct'] = False
        
    elif '5' in opciones_seleccionadas:  # Opciones de correo
        print("\n🚀 Procesando métodos de envío por correo...")
        
        # 1. Enviar resumen por correo (HTML)
        print("\n📧 Enviando resumen por correo (HTML)...")
        resultados['email_html'] = enviar_pedido_por_correo(resumen_pedido, datos_cliente)
        
        # 2. Generar y enviar factura Excel
        print("\n📊 Generando y enviando factura Excel...")
        exito_gen_factura, exito_envio, ruta_factura = generar_y_enviar_factura(datos_cliente, resumen_pedido)
        resultados['email_excel'] = exito_envio
        
    else:
        # Procesar opciones individuales
        if '1' in opciones_seleccionadas:
            print("\n📧 Enviando resumen por correo (HTML)...")
            resultados['email_html'] = enviar_pedido_por_correo(resumen_pedido, datos_cliente)
            
        if '2' in opciones_seleccionadas:
            print("\n📊 Generando y enviando factura Excel...")
            exito_gen_factura, exito_envio, ruta_factura = generar_y_enviar_factura(datos_cliente, resumen_pedido)
            resultados['email_excel'] = exito_envio
            
        if '3' in opciones_seleccionadas:
            print("\n📱 Enviando mensaje por WhatsApp (Selenium)...")
            resultados['whatsapp'] = enviar_pedido_por_whatsapp(datos_cliente['telefono'], resumen_pedido, datos_cliente)
            
            # Si ya hemos generado la factura, preguntar si quiere enviarla por WhatsApp también
            if exito_gen_factura and ruta_factura:
                enviar_factura = input("\n¿Deseas enviar también la factura Excel por WhatsApp? (s/n): ").strip().lower()
                if enviar_factura.startswith('s'):
                    print("\n📎 Enviando factura Excel por WhatsApp...")
                    resultados['whatsapp_file'] = enviar_factura_por_whatsapp(
                        datos_cliente['telefono'], 
                        ruta_factura, 
                        datos_cliente, 
                        resumen_pedido
                    )
            
        if '4' in opciones_seleccionadas:
            print("\n📱 Generando enlace y QR para WhatsApp...")
            try:
                # Verificar si la biblioteca qrcode está disponible
                resultado_direct = enviar_pedido_whatsapp_directo(
                    datos_cliente['telefono'], 
                    resumen_pedido, 
                    datos_cliente
                )
                resultados['whatsapp_direct'] = True
            except NameError:
                print("❌ Error: Se requiere instalar la biblioteca 'qrcode'. Ejecuta: pip install qrcode[pil]")
                resultados['whatsapp_direct'] = False
            except Exception as e:
                print(f"❌ Error al generar enlace de WhatsApp: {e}")
                resultados['whatsapp_direct'] = False
    
    # Mostrar resumen de resultados
    print("\n📋 RESUMEN DE ENVÍOS:")
    if resultados['email_html']:
        print("✅ Resumen HTML enviado correctamente por correo")
    else:
        print("❌ No se envió el resumen HTML por correo")
        
    if resultados['email_excel']:
        print("✅ Factura Excel enviada correctamente por correo")
    else:
        print("❌ No se envió la factura Excel por correo")
        
    if resultados['whatsapp']:
        print("✅ Mensaje enviado correctamente por WhatsApp (Selenium)")
    else:
        if '3' in opciones_seleccionadas or '6' in opciones_seleccionadas:
            print("❌ No se pudo enviar mensaje por WhatsApp usando Selenium")
    
    if resultados['whatsapp_file']:
        print("✅ Factura Excel enviada correctamente por WhatsApp")
    else:
        if resultados['whatsapp'] and (exito_gen_factura and ruta_factura):
            print("❌ No se pudo enviar la factura por WhatsApp")
        
    if resultados['whatsapp_direct']:
        print("✅ Enlace y código QR para WhatsApp generados correctamente")
    
    return any(resultados.values())

def test_whatsapp_direct():
    """Función de prueba para enviar WhatsApp directamente a un número específico"""
    numero = "3042535003"  # Tu número
    print(f"\n=== PRUEBA DE WHATSAPP PARA NÚMERO {numero} ===\n")
    
    mensaje = (
        f"🍔 *PRUEBA DE SAMIR'S BURGERS* 🍔\n\n"
        f"Este es un mensaje de prueba enviado desde el sistema de Samir's Burgers.\n\n"
        f"Si estás recibiendo esto, ¡la configuración está funcionando correctamente!\n\n"
        f"Hora del envío: {datetime.datetime.now().strftime('%H:%M:%S')}"
    )
    
    print("Generando enlace directo y QR para WhatsApp...")
    
    try:
        if QR_DISPONIBLE:
            resultado = enviar_pedido_whatsapp_directo(numero, mensaje, {"nombre": "Cliente de prueba", "direccion": "Dirección de prueba", "metodo_pago": "Efectivo"})
            return resultado['browser_opened']
        else:
            print("Para generar QR, instala: pip install qrcode[pil]")
            enlace = crear_enlace_whatsapp(numero, mensaje)
            print(f"Enlace directo: {enlace}")
            return abrir_chat_directo(numero, mensaje)
    except Exception as e:
        print(f"Error en prueba de WhatsApp: {e}")
        return False

def main():
    """
    Flujo del chatbot:
      1. El usuario conversa con el asistente y realiza su pedido.
      2. Al escribir "confirmar", se finaliza el pedido.
      3. Se solicitan de forma individual los datos de contacto del cliente.
      4. Se genera y muestra el resumen final del pedido.
      5. Se presentan opciones de envío: correo HTML, factura Excel y WhatsApp.
    """
    conversacion = [
        {
            "role": "system",
            "content": (
                "Eres un asistente de Samir's Burgers en Medellín. "
                "Nuestro menú incluye: Hamburguesa Clásica (6000), Hamburguesa Doble (8000), Papas Fritas (3000). "
                "El domicilio cuesta 2000, pero es gratis en pedidos mayores a 15000. "
                "No solicites aún la dirección, nombre, teléfono y método de pago; se agregarán al finalizar el pedido, "
                "cuando el usuario escriba 'confirmar'. Luego, se mostrará el resumen final y las opciones de envío."
            )
        },
        {
            "role": "user",
            "content": "¿Qué precios, combos y promociones tienes?"
        }
    ]
    
    print("\n" + "="*60)
    print("🍔 Bienvenido a 'Samir's Burgers' 🍔")
    print("="*60 + "\n")
    print("¿En qué puedo ayudarte hoy?")
    print("¿Qué precios, combos y promociones tienes?")
    print("¿Qué deseas ordenar hoy?\n")
    
    # Bucle de conversación de pedido
    while True:
        user_input = input("Usuario: ")
        user_input_norm = user_input.strip().lower()
        
        if user_input_norm == "confirmar":
            print("\n✅ Has finalizado tu pedido correctamente.")
            break
        
        if user_input_norm == "salir" or user_input_norm == "exit":
            print("\n👋 Gracias por visitarnos. ¡Hasta pronto!")
            sys.exit()
        
        if user_input_norm == "enviar":
            print("⚠️ No se puede enviar aún. Primero debes confirmar tu pedido escribiendo 'confirmar'.")
            continue
        
        conversacion.append({"role": "user", "content": user_input})
        respuesta = obtener_respuesta(conversacion)
        conversacion.append({"role": "assistant", "content": respuesta})
        print(f"Asistente: {respuesta}\n")
    
    # Solicitar datos de contacto de forma individual
    print("\n📋 Para completar tu pedido, por favor responde lo siguiente:")
    direccion = input("¿Cuál es la dirección de entrega? ")
    nombre_cliente = input("¿Cuál es tu nombre? ")
    telefono = input("¿Cuál es tu número de teléfono? ")
    
    # Opciones de método de pago
    print("\n💰 Selecciona tu método de pago:")
    print("1. Efectivo")
    print("2. Transferencia/Nequi")
    print("3. Tarjeta de crédito/débito (al recibir)")
    
    metodo_pago = ""
    while not metodo_pago:
        opcion = input("Ingresa el número de tu método de pago (1-3): ").strip()
        if opcion == "1":
            metodo_pago = "Efectivo"
        elif opcion == "2":
            metodo_pago = "Transferencia/Nequi"
        elif opcion == "3":
            metodo_pago = "Tarjeta (al recibir)"
        else:
            print("❌ Opción no válida. Intenta nuevamente.")
    
    # Datos del cliente
    datos_cliente = {
        "direccion": direccion,
        "nombre": nombre_cliente,
        "telefono": telefono,
        "metodo_pago": metodo_pago
    }
    
    # Generar resumen del pedido
    print("\n⏳ Generando resumen del pedido...")
    resumen_pedido = generar_resumen_pedido(conversacion, datos_cliente)
    
    # Mostrar resumen final
    print("\n" + "="*60)
    print("📝 RESUMEN DE TU PEDIDO:")
    print("="*60)
    print(resumen_pedido)
    print("="*60)
    
    # Procesar opciones de envío
    procesar_opciones_envio(datos_cliente, resumen_pedido)
    
    print("\n🎉 ¡Pedido completado con éxito! 🎉")
    print("Gracias por ordenar en Samir's Burgers. ¡Buen provecho!")

def test_mode():
    """Modo de prueba para probar componentes individuales"""
    print("\n" + "="*60)
    print("🧪 MODO DE PRUEBA - Samir's Burgers 🧪")
    print("="*60 + "\n")
    
    print("Selecciona una opción de prueba:")
    print("1. Probar envío de mensaje WhatsApp (Selenium)")
    print("2. Probar WhatsApp directo (enlace/QR)")
    print("3. Probar generación de factura Excel")
    print("4. Probar envío de correo")
    print("5. Salir")
    
    opcion = input("\nIngresa el número de opción: ").strip()
    
    if opcion == "1":
        # Probar WhatsApp con Selenium
        numero = input("Ingresa un número de teléfono para la prueba: ")
        mensaje = "🍔 Este es un mensaje de prueba desde Samir's Burgers"
        print(f"Enviando mensaje de prueba a {numero}...")
        
        bot = WhatsAppBot()
        if bot.start():
            result = bot.send_message(numero, mensaje)
            bot.close()
            
            if result:
                print("✅ Mensaje enviado correctamente")
            else:
                print("❌ No se pudo enviar el mensaje")
        else:
            print("❌ No se pudo iniciar WhatsApp Web")
            
    elif opcion == "2":
        # Probar WhatsApp directo
        test_whatsapp_direct()
        
    elif opcion == "3":
        # Probar generación de factura
        datos_prueba = {
            "nombre": "Cliente de Prueba",
            "telefono": "3001234567",
            "direccion": "Calle de Prueba #123",
            "metodo_pago": "Efectivo"
        }
        
        resumen_prueba = (
            "1 Hamburguesa Clásica - $6000\n"
            "1 Papas Fritas - $3000\n"
            "1 Gaseosa - $2000\n"
            "Domicilio - $2000"
        )
        
        factura_gen = FacturaExcel()
        ruta = factura_gen.generar_factura(datos_prueba, resumen_prueba)
        
        print(f"✅ Factura generada correctamente: {ruta}")
        
    elif opcion == "4":
        # Probar envío de correo
        datos_prueba = {
            "nombre": "Cliente de Prueba",
            "telefono": "3001234567",
            "direccion": "Calle de Prueba #123",
            "metodo_pago": "Efectivo"
        }
        
        resumen_prueba = (
            "1 Hamburguesa Clásica - $6000\n"
            "1 Papas Fritas - $3000\n"
            "1 Gaseosa - $2000\n"
            "Domicilio - $2000"
        )
        
        resultado = enviar_pedido_por_correo(resumen_prueba, datos_prueba)
        
        if resultado:
            print("✅ Correo enviado correctamente")
        else:
            print("❌ No se pudo enviar el correo")
            
    elif opcion == "5":
        print("Saliendo del modo de prueba...")
        return
    
    else:
        print("❌ Opción no válida")
    
    # Preguntar si desea realizar otra prueba
    otra_prueba = input("\n¿Deseas realizar otra prueba? (s/n): ").strip().lower()
    if otra_prueba.startswith('s'):
        test_mode()

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        test_mode()
    else:
        main()